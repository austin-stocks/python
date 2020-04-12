
import csv
import openpyxl
from openpyxl.styles import PatternFill
import os
import xlrd
import sys
import time
import pandas as pd
import datetime as dt
from yahoofinancials import YahooFinancials
import time
import logging
import xlsxwriter


#
# Define the directories and the paths
dir_path = os.getcwd()
user_dir = "\\..\\" + "User_Files"
chart_dir = "..\\" + "Charts"
historical_dir = "\\..\\" + "Historical"
earnings_dir = "\\..\\" + "Earnings"
dividend_dir = "\\..\\" + "Dividend"
log_dir = "\\..\\" + "Logs"
analysis_dir = "\\..\\" + "Analysis"
# ---------------------------------------------------------------------------
# Set Logging
# critical, error, warning, info, debug
# set up logging to file - see previous section for more details
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s',
                    datefmt='%m-%d %H:%M',
                    filename=dir_path + log_dir + "\\" + 'SC_Parese_AAII_Download_debug.txt',
                    filemode='w')
# define a Handler which writes INFO messages or higher to the sys.stderr
console = logging.StreamHandler()
console.setLevel(logging.INFO)
# set a format which is simpler for console use
formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
# tell the handler to use this format
console.setFormatter(formatter)
# add the handler to the root logger
logging.getLogger('').addHandler(console)

# Disnable and enable global level logging
logging.disable(sys.maxsize)
logging.disable(logging.NOTSET)
# ---------------------------------------------------------------------------



lynch_analysis_start_row_number = 3
lynch_analysis_stop_row_number = 6

phil_town_analysis_stop_row_number = 20

tracklist_file = "Tracklist.csv"
tracklist_file_full_path = dir_path + user_dir + "\\" + tracklist_file
configuration_file = "Configurations.csv"
configurations_file_full_path = dir_path + user_dir + "\\" + configuration_file
tracklist_df = pd.read_csv(tracklist_file_full_path)





# -----------------------------------------------------------------------------
# Read the AAII file
# -----------------------------------------------------------------------------
# This takes around 21 sec
start = time.process_time()
aaii_xls = pd.ExcelFile('2020_04_01_AAII_Analysis.xlsx')
print(time.process_time() - start)

qtr_str_list =['Q1', 'Q2','Q3','Q4','Q5','Q6','Q7','Q8']
yr_str_list =['Y1', 'Y2','Y3','Y4','Y5','Y6','Y7']

aaii_dateandperiod_df = pd.read_excel(aaii_xls, 'Dates')
aaii_bs_qtr_df = pd.read_excel(aaii_xls, 'Balance_QTR')
aaii_pnl_qtr_df  = pd.read_excel(aaii_xls, 'Income_QTR')
aaii_bs_yr_df = pd.read_excel(aaii_xls, 'Balance_YR')
aaii_pnl_yr_df  = pd.read_excel(aaii_xls, 'Income_YR')

# Set the Ticker col and index
aaii_dateandperiod_df.set_index('Ticker', inplace=True)
aaii_bs_qtr_df.set_index('Ticker', inplace=True)
aaii_pnl_qtr_df.set_index('Ticker', inplace=True)
aaii_bs_yr_df.set_index('Ticker', inplace=True)
aaii_pnl_yr_df.set_index('Ticker', inplace=True)
# -----------------------------------------------------------------------------


ticker_list_unclean = tracklist_df['Tickers'].tolist()
ticker_list = [x for x in ticker_list_unclean if str(x) != 'nan']

# #############################################################################
# #############################################################################
# #############################################################################
#                   MAIN LOOP FOR TICKERS
# #############################################################################
# #############################################################################
# #############################################################################
# It seems that there will be 3 parts of the code
# 1. To read in the AAII data and prepare a dataframe that has the Analysis
#   1a. We can create one dataframe per "Style of Analysis" - like one for lynch, one for Phil Town etc
# 2. To Read in the existing analysis file and "merge" the existing analysis data with the newly created Analysis data from AAII
#   2a. Break that data into various dataframes each representing a styly - lynch, Phil town etc
# 3. Write that data back into the analysis file

# ticker_list = ['AAPL', 'AUDC','MED']
for ticker_raw in ticker_list:

  ticker = ticker_raw.replace(" ", "").upper() # Remove all spaces from ticker_raw and convert to uppercase
  logging.info("========================================================")
  logging.info("Processing for " + ticker)
  logging.info("========================================================")

  # Get the various ticker information from the various dfs
  aaii_dateandperioed_series = aaii_dateandperiod_df.loc[ticker]
  aaii_bs_qtr_series = aaii_bs_qtr_df.loc[ticker]
  aaii_pnl_qtr_series = aaii_pnl_qtr_df.loc[ticker]
  aaii_bs_yr_series = aaii_bs_yr_df.loc[ticker]
  aaii_pnl_yr_series = aaii_pnl_yr_df.loc[ticker]

  logging.debug("The date and period series is : " + str(aaii_dateandperioed_series))
  logging.debug("The Balance Sheet YR series is : " + str(aaii_bs_yr_series))
  logging.debug("The Balance Sheet QTR series is : " + str(aaii_bs_qtr_series))


  # ---------------------------------------------------------------------------
  # Get the various fields that we need for Analysis and prepare the analysis df
  # ---------------------------------------------------------------------------
  aaii_qtr_dates_dict_dt = {}
  aaii_qtr_dates_dict_str = {}
  aaii_yr_dates_dict_dt = {}
  aaii_yr_dates_dict_str = {}

  for qtr_idx in qtr_str_list:
    logging.debug("Getting the Quarterly data from AAII for " + str(qtr_idx))
    aaii_qtr_dates_dict_dt[qtr_idx] =  dt.datetime.strptime(str(aaii_dateandperioed_series['Ending date ' + str(qtr_idx)]),'%Y-%m-%d %H:%M:%S').date()
    aaii_qtr_dates_dict_str[qtr_idx] =  aaii_qtr_dates_dict_dt[qtr_idx].strftime('%m/%d/%Y')

  aaii_yr_df = pd.DataFrame(columns=['AAII_YR_DATA'])
  aaii_yr_df.set_index(['AAII_YR_DATA'], inplace=True)
  for yr_idx in yr_str_list:
    logging.debug("Getting the Yearly data from AAII for " + str(yr_idx))
    aaii_yr_dates_dict_dt[yr_idx] = dt.datetime.strptime(str(aaii_dateandperioed_series['Ending date ' + str(yr_idx)]),'%Y-%m-%d %H:%M:%S').date()
    aaii_yr_dates_dict_str[yr_idx] =  aaii_yr_dates_dict_dt[yr_idx].strftime('%m/%d/%Y')
    tmp_val = aaii_yr_dates_dict_str[yr_idx]
    aaii_yr_df.assign(tmp_val = "")
    aaii_yr_df.loc['Revenue', tmp_val] = aaii_pnl_yr_series['Sales '+str(yr_idx)]
    aaii_yr_df.loc['Dividend', tmp_val] = aaii_pnl_yr_series['Dividend '+str(yr_idx)]
    aaii_yr_df.loc['Diluted_EPS', tmp_val] = aaii_pnl_yr_series['EPS-Diluted Continuing '+str(yr_idx)]
    aaii_yr_df.loc['LT_Debt', tmp_val] = aaii_bs_yr_series['Long-term debt '+str(yr_idx)]
    aaii_yr_df.loc['BV_Per_Share', tmp_val] = aaii_bs_yr_series['Book value/share '+str(yr_idx)]


  logging.debug("\n\nThe YR DF Prepared from AAII Data is \n" + aaii_yr_df.to_string() + "\n")
  aaii_yr_df_col_list = aaii_yr_df.columns.tolist()
  logging.debug("The column list for AAII YR DF is " + str(aaii_yr_df_col_list))
  aaii_yr_date_list_dt = [dt.datetime.strptime(date, '%m/%d/%Y').date() for date in aaii_yr_df_col_list]
  logging.debug("\n\nThe AAII YR DF Datelist is " + str(aaii_yr_date_list_dt) + " and the number of elements is " + str(len(aaii_yr_date_list_dt)))
  # ---------------------------------------------------------------------------


  # ---------------------------------------------------------------------------
  # Now generate the aaii_lynch_qtr and aaii_aaii_lynch_yr_df
  # ---------------------------------------------------------------------------
  aaii_lynch_yr_df = aaii_yr_df.copy()
  # Rename the index
  aaii_lynch_yr_df.rename_axis("AAII_LYNCH_YR_DATA", inplace=True)

  # Drop the rows from dataframe that are not needed for lynch analysis
  aaii_lynch_yr_df.drop(['Dividend'], axis=0, inplace=True)
  # logging.debug("The YR_Lynch_df is \n" + aaii_lynch_yr_df.to_string())

  # Re-sort the df based on column names - by default it is sorted in ascending order
  aaii_lynch_yr_df_col_list = aaii_lynch_yr_df.columns.tolist()
  aaii_lynch_yr_df = aaii_lynch_yr_df.reindex(sorted(aaii_lynch_yr_df.columns), axis=1)
  logging.debug("\n\nThe sorted YR_Lynch_df by the dates : \n" + aaii_lynch_yr_df.to_string() + "\n")
  # ---------------------------------------------------------------------------





  # ###########################################################################
  #                PART 2 - Read the alreay existing Analysis Data
  #                   and merge it with created AAII Analysis Data
  # ###########################################################################
  ticker_xls_exists = 0
  if (os.path.exists(dir_path + analysis_dir + "\\" + ticker + "_yr_Analysis.xlsx") is True):
    logging.debug("The Analysis file exists")
    ticker_xls = pd.ExcelFile(dir_path + analysis_dir + "\\" + ticker + "_yr_Analysis.xlsx")
    ticker_yr_analysis_df = pd.read_excel(ticker_xls, 'Yearly', index_col=0)
    ticker_xls_exists = 1

  ticker_lynch_df_start_row = 3
  ticker_lynch_df_stop_row = 4
  ticker_phil_town_df_start_row = 7
  ticker_phil_town_df_stop_row = 9

  # If the Analysis file exits for the ticker then merge that date - for each analysis style with the created data
  if (ticker_xls_exists == 1):
    # ticker_yr_analysis_df.set_index([0])
    logging.debug("\n\nThe Ticker Yearly Analysis df is \n" + ticker_yr_analysis_df.to_string() + "\n")
    # Convert the timestamp into datetime
    ticker_yr_analysis_date_list_dt = [x.date() for x in list(ticker_yr_analysis_df)]
    logging.debug("The datelist dt for ticker yr analysis df is " + str(ticker_yr_analysis_date_list_dt))

    # Now split the dataframe to get the ticker_lynch_yr_dataframe
    ticker_lynch_yr_df = ticker_yr_analysis_df[ticker_lynch_df_start_row-2:ticker_lynch_df_stop_row-1]
    ticker_phil_town_yr_df = ticker_yr_analysis_df[ticker_phil_town_df_start_row-2:ticker_phil_town_df_stop_row-1]
    ticker_lynch_yr_df.rename_axis("Lynch_Analysis", inplace=True)
    logging.debug("\n\nThe ticker Lynch df is \n" + ticker_lynch_yr_df.to_string() + "\n")

    # Make row0 as the column headers
    # ticker_phil_town_yr_df.rename(columns=ticker_phil_town_yr_df.iloc[1])
    # ticker_phil_town_yr_df.columns = ticker_phil_town_yr_df.iloc[0]
    ticker_phil_town_yr_df.drop(ticker_phil_town_yr_df.index[0])
    # ticker_phil_town_yr_df.drop(['Phil_Town_Analysis'], axis=0, inplace=True)
    logging.debug("\n\nThe ticker Phil Town df is \n" + ticker_phil_town_yr_df.to_string() + "\n")


    # Now we have two date_dt list that need to be compared
    logging.debug("\n\nThe AAII YR DF Datelist is " + str(type(aaii_yr_date_list_dt)) + " \nand the number of elements is " + str(len(aaii_yr_date_list_dt)))
    for date_dt_idx in aaii_yr_date_list_dt:
      if (date_dt_idx in ticker_yr_analysis_date_list_dt):
        logging.debug(str(date_dt_idx) + " in AAII df is present in ticker df. Will need to replace...")
      else:
        logging.debug(str(date_dt_idx) + " in AAII df is not present in ticker df. Will need to insert...")


  logging.debug("All Done")
  sys.exit(1)


  tmp_list = ticker_yr_analysis_df.columns.tolist()
  ticker_yr_analysis_col_list = [dt.datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S').date() for date in tmp_list]
  logging.debug("The type of tmp_list is " + str(type(tmp_list)))
  logging.debug("The type of xxx list is " + str(type(ticker_yr_analysis_col_list)))
  print("The type of tmp_list is ",  type(tmp_list))
  print("The type of xxx list is ", type(ticker_yr_analysis_col_list[0]))

  for date_idx in aaii_lynch_yr_df_col_list:
    logging.debug("Checking for date " + str(date_idx) + " in the ticker column list " + str(ticker_yr_analysis_col_list))
    if date_idx in ticker_yr_analysis_col_list:
      logging.debug("Found it")
    else:
      logging.debug("Adding a column for " + str(date_idx))
      print("The value of date_idx ", date_idx , " and the type ", type(date_idx))
      # ticker_yr_analysis_df.assign(date_idx="")
      ticker_yr_analysis_df.insert(0,str(date_idx),"Haha")
      ticker_yr_analysis_df[str(date_idx)] = aaii_lynch_yr_df[date_idx]

  aaii_lynch_yr_df = aaii_lynch_yr_df.reindex(sorted(aaii_lynch_yr_df.columns), axis=1)
  logging.debug("The ticker df now is \n" + ticker_yr_analysis_df.to_string())

  # for i_idx in range(1,9):
  #   qtr_str = "Q"+str(i_idx)
  #   tmp_val = aaii_qtr_dates_dict_dt[qtr_str]
  #   lynch_qtr_df.assign(tmp_val= "")
  #
  # lynch_qtr_df.set_index(lynch_qtr_df_index, inplace=True)
  # logging.debug("The QTR_Lynch_df is \n" + lynch_qtr_df.to_string())
  #
  # for i_idx in range(1,9):
  #   qtr_str = "Q"+str(i_idx)
  #   logging.debug("The qtr_str is " + str(qtr_str) + " and the date value is  " + str(aaii_qtr_dates_dict_dt[qtr_str]) + " and the revenue is " + str(ticker_qtr_revenue_list[i_idx-1]))
  #   lynch_qtr_df.loc['Revenue', aaii_qtr_dates_dict_dt[qtr_str].strftime('%m/%d/%Y')] = ticker_qtr_revenue_list[i_idx-1]
  #
  # logging.debug("The QTR_Lynch_df is \n" + lynch_qtr_df.to_string())


sys.exit(1)

# -----------------------------------------------------------------------------
# The program write a combination of dataframe and random data
# -----------------------------------------------------------------------------
# Create an pandas excel writer based on xlsxwriter.
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
workbook  = writer.book

df1 = pd.DataFrame({'Data': [11, 12, 13, 14]})
df2 = pd.DataFrame({'Data': [21, 22, 23, 24]})
df3 = pd.DataFrame({'Data': [31, 32, 33, 34]})
df4 = pd.DataFrame({'Data': [41, 42, 43, 44]})


# Peter Lynch Numbers
# 1. Revenue/Sales growing year-over-year and qtr-over-qtr - Income Statement
# 2. P/E Ratio Relative to Growth - Need to think about it (Need to think about it)
# 3. Cash net of long term debt - Balance Sheet
# 4. Debt/Equity - Balance Sheet
# 5. Dividends - Yahoo Finance or some other sheet in AAII but in the charts anyway
# 6. Does it Pay? - Not sure
# 7. Book Value - Balance Sheet
# 8. More Hidden assets - Not sure
# 9. Cash Flow - Income and Cash Flow
# 10. Inventories - Balance Sheet
# 11. Pension Plans - Not sure
# 12. Growth Rate of Earnings - In the chart (the growth curves)
# 13. The Bottom Line - Not sure


df5_row = "Lynch_Analysis"
df5_cols = [df5_row,'W','X','Y','Z']
df5 = pd.DataFrame(columns=df5_cols)
df5.set_index(df5_row,inplace=True)
df5.loc['Cash_Position_Change', 'W'] = "Something_Random"

logging.debug("The df5 is \n" + df5.to_string())





df1.to_excel(writer, sheet_name='Sheet1')  # Default position, cell A1.
df2.to_excel(writer, sheet_name='Sheet1', startcol=3)
df3.to_excel(writer, sheet_name='Sheet1', startrow=6)

# It is also possible to write the dataframe without the header and index.
df4.to_excel(writer, sheet_name='Sheet1',
             startrow=7, startcol=4, header=False, index=False)

worksheet = writer.sheets['Sheet1']
# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})
# Add a number format for cells with money.
money = workbook.add_format({'num_format': '$#,##0'})

# Write some data headers.
worksheet.write('A1', 'Item', bold)
worksheet.write('B1', 'Cost', bold)

# Some data we want to write to the worksheet.
expenses = (
 ['Rent', 1000],
 ['Gas',   100],
 ['Food',  300],
 ['Gym',    50],
)

# Start from the first cell below the headers.
row = 1
col = 0

# Iterate over the data and write it out row by row.
for item, cost in (expenses):
 worksheet.write(row, col,     item)
 worksheet.write(row, col + 1, cost, money)
 row += 1

# Write a total using a formula.
worksheet.write(row, 0, 'Total',       bold)
worksheet.write(row, 1, '=SUM(B2:B5)', money)

# Access the underlying xlsxwriter worksheet and write to it.
worksheet.write(0, 0, 'test')

# Apply conditional formatting
format1 = workbook.add_format({'bg_color':   '#FFC7CE',
                               'font_color': '#9C0006'})
worksheet.conditional_format('A1:Z1024', {'type':     'cell',
                                          'criteria': '>=',
                                          'value':    50,
                                          'format':   format1})

writer.save()

# Now save that as an image
import excel2img

# Save as PNG the range of used cells in test.xlsx on page named "Sheet1"
# excel2img.export_img("test.xlsx", "test.png", "Sheet1", None)

# Save as BMP the range B2:C15 in test.xlsx on page named "Sheet2"
excel2img.export_img("test.xlsx", "test.png", "", "Sheet1!A1:J30")


print("All Done")
# -----------------------------------------------------------------------------




# for col_range in range(1, 12):
#     cell_title = sheet.cell(1, col_range)
#     cell_title.fill = PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")
#
'''
# This takes around 141 sec
start = time.process_time()
xls = pd.ExcelFile('2019_04_01_AAII_DATA_Experiment.xlsm')
print(time.process_time() - start)

# This takes around 11 sec
start = time.process_time()
aaii_dateandperiod_df = pd.read_excel(xls, 'DateAndPeriod')
aaii_companyinfo_df = pd.read_excel(xls, 'CompanyInformation')
aaii_income_qtr_df = pd.read_excel(xls, 'IncomeSheet-QTR')
aaii_income_yr_df = pd.read_excel(xls, 'IncomeSheet-YR')
aaii_balancesheet_qtr_df = pd.read_excel(xls, 'BalanceSheet-QTR')
aaii_balancesheet_yr_df = pd.read_excel(xls, 'BalanceSheet-YR')
aaii_cashflow_qtr_df = pd.read_excel(xls, 'CashFlow-QTR')
aaii_cashflow_yr_df = pd.read_excel(xls, 'CashFlow-YR')
aaii_estimates_df = pd.read_excel(xls, 'EstimatesAndMultiples')
aaii_growth_df = pd.read_excel(xls, 'EstimatesAndMultiples')
aaii_misc_qtr_df = pd.read_excel(xls, 'Misc-QTR')
aaii_misc_yr_df = pd.read_excel(xls, 'Misc-YR')
aaii_rank_df = pd.read_excel(xls, 'Rank')
aaii_ratio_df = pd.read_excel(xls, 'RatiosAndValuations')
aaii_sector_df = pd.read_excel(xls, 'Median-Sectors')
aaii_industry_df = pd.read_excel(xls, 'Median-Industries')
print(time.process_time() - start)


ticker_list_unclean = aaii_dateandperiod_df['Ticker'].tolist()
ticker_list = [x for x in ticker_list_unclean if str(x) != 'nan']

for ticker_raw in ticker_list:

  ticker = ticker_raw.replace(" ", "").upper() # Remove all spaces from ticker_raw and convert to uppercase
  logging.info("========================================================")
  logging.info("Preparing 0 Ananlysis for " + ticker)
  logging.info("========================================================")

print ("All Done...")
'''


#
# # 6/30/2019 - This scripte works as inteneded but get thrown off when there are Yahoo and CNBC projections for the
# #  same date - an example is AAP - look at date 4/21/2017 this has Yahoo and CNBC data (In my opinion incorrectly
# # as Yahoo date is not the same row as the earnings is). Anyway because of the way the script is written it will
# # put the Yahoo date in the previous projections as all the rows what have nan in projection are deleted before
# # the processing so currently the scirpt has not way of knowing whether the Yahoo data is for the previous quater
# # or for the current quater. This shows up in the Earnings.csv file for AAP
# # One solution is to delted all the rows that have Yahoo is the column immpediately following projection column
# # that way we will lose that data but that should not be a big deal
#
# tracklist_df = pd.read_csv('Get_Earnings_Tracklist.csv')
# ticker_list_unclean = tracklist_df['Tickers'].tolist()
# ticker_list = [x for x in ticker_list_unclean if str(x) != 'nan']
#
# for ticker_raw in ticker_list:
#   ticker = ticker_raw.replace(" ", "").upper()  # Remove all spaces from ticker_raw and convert to uppercase
#   print("Getting Earnings for ", ticker)
#
#   # todo : Why does this not work here?
#   # earnings_df = pd.read_excel(dir_path + "\\" + "Stock_Files" + "\\" + ticker + '.xlsm', sheet_name="historical")
#   earnings_df = pd.read_excel('C:\Sundeep\Stocks_Automation\Scripts\Experiments\Stocks_Files' + "\\" + ticker + '.xlsm', sheet_name="historical")
#
#   print ("The Historical Tab (which contains earnings data) from the stock file is :", earnings_df)
#   print ("\nAll the columns in the historical tab are :\n")
#   for col in earnings_df.columns:
#     print(col)
#
#   # For some reason python reads the Date column as Timestamp
#   # Convert it to string here We will convet it to Datetime later
#   date_list = earnings_df['Date'].astype(str).tolist()
#   qtr_eps_list = earnings_df['Q EPS'].tolist()
#   projected_eps_list = earnings_df['projection'].tolist()
#   print ("The Raw Date list is :", date_list)
#
#   # Check if the length of all the columns are equal
#
#   # This works : The create a dataframe from a list of lists
#   step1_df=pd.DataFrame(list(zip(date_list, qtr_eps_list, projected_eps_list)),
#                         columns=['Date','Q EPS', 'projection'])
#   print ("Dataframe that only has Date, Q EPS and projection columns :", step1_df)
#
#   # ===========================================================================
#   # Clean up the dataframe
#   # ===========================================================================
#   # Remove all rows that do not have data in ALL columns
#   step1_df.dropna(how='all',inplace=True)
#   # Remove all rows that do not have data in Date column
#   step1_df.dropna(subset=['Date'],inplace=True)
#   tmp_df = step1_df[step1_df.Date != 'NaT']
#   step1_df = tmp_df
#   print ("New dataframe after dropping all rows with null in Date column:", step1_df)
#   # step1_df.to_csv('debug1.csv')
#   # Now we should have a dataframe that does not have any rows that do not have a valid date
#
#   # Drop ONLY the rows now that do not have any data in Q EPA AND projection column
#   step1_df.dropna(subset=['Q EPS', 'projection'], how='all', inplace=True)
#   # step1_df.to_csv('debug.csv')
#   print ("New dataframe after dropping all the rows with null in (Q EPS AND projection) columns:", step1_df)
#   # ===========================================================================
#
#   # ===========================================================================
#   # Extract the lists and write csv file
#   # ===========================================================================
#   csvFile=open('Extracted_Earnings' + "\\" + ticker + "_earnings.csv", 'w+', newline='')
#   writer = csv.writer(csvFile)
#   # Put the Header Row in the csv
#   writer.writerow(["Date", "Q_EPS_Diluted"])
#
#   step1_date_list = [dt.datetime.strptime(date, '%Y-%m-%d').date() for date in step1_df.Date.tolist()]
#   # step1_date_list = [dt.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').date() for date in step1_df.Date.tolist()]
#   step1_eps_list = step1_df['Q EPS'].tolist()
#   step1_projected_eps_list = step1_df['projection'].tolist()
#
#   tmp_index = 0
#   csv_line = []
#   for x in step1_date_list:
#     step1_eps = step1_eps_list[tmp_index]
#     step1_projected_eps = step1_projected_eps_list[tmp_index]
#     print ("Index : ", tmp_index, ", Date is : ", x, ", Q EPS is : ", step1_eps, ", Projected EPS is : ", step1_projected_eps)
#     if (str(step1_eps) != 'nan'):
#       print ("CSV String is :", csv_line)
#       if (tmp_index > 0):
#         writer.writerow(csv_line)
#
#       csv_line = []
#       print ("Found 1st Instance")
#       csv_line.insert(0, x.strftime('%m/%d/%Y'))
#       csv_line.insert(1, step1_eps)
#       csv_line.insert(2, step1_projected_eps)
#       insert_index = 3
#     else:
#       csv_line.insert(insert_index, step1_projected_eps)
#       insert_index = insert_index+1
#     tmp_index = tmp_index+1
#
# csvFile.close()
#
