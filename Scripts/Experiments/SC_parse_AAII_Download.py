

import csv
import openpyxl
from openpyxl.styles import PatternFill
import os
import xlrd
import sys
import time
import pandas as pd
import datetime as dt
from yahoofinancials import YahooFinancials
import time
import logging
import xlsxwriter

# ---------------------------------------------------------------------------
# Set Logging
# critical, error, warning, info, debug
# set up logging to file - see previous section for more details
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s',
                    datefmt='%m-%d %H:%M',
                    filename='Logging_AAII_Parser.txt',
                    filemode='w')
# define a Handler which writes INFO messages or higher to the sys.stderr
console = logging.StreamHandler()
console.setLevel(logging.INFO)
# set a format which is simpler for console use
formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
# tell the handler to use this format
console.setFormatter(formatter)
# add the handler to the root logger
logging.getLogger('').addHandler(console)

# Disnable and enable global level logging
logging.disable(sys.maxsize)
logging.disable(logging.NOTSET)
# ---------------------------------------------------------------------------

# your code here
dir_path = os.getcwd()

# -----------------------------------------------------------------------------
# The program write a combination of dataframe and random data
# -----------------------------------------------------------------------------
# Create an pandas excel writer based on xlsxwriter.
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
workbook  = writer.book

df1 = pd.DataFrame({'Data': [11, 12, 13, 14]})
df2 = pd.DataFrame({'Data': [21, 22, 23, 24]})
df3 = pd.DataFrame({'Data': [31, 32, 33, 34]})
df4 = pd.DataFrame({'Data': [41, 42, 43, 44]})

df1.to_excel(writer, sheet_name='Sheet1')  # Default position, cell A1.
df2.to_excel(writer, sheet_name='Sheet1', startcol=3)
df3.to_excel(writer, sheet_name='Sheet1', startrow=6)

# It is also possible to write the dataframe without the header and index.
df4.to_excel(writer, sheet_name='Sheet1',
             startrow=7, startcol=4, header=False, index=False)

worksheet = writer.sheets['Sheet1']
# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})
# Add a number format for cells with money.
money = workbook.add_format({'num_format': '$#,##0'})

# Write some data headers.
worksheet.write('A1', 'Item', bold)
worksheet.write('B1', 'Cost', bold)

# Some data we want to write to the worksheet.
expenses = (
 ['Rent', 1000],
 ['Gas',   100],
 ['Food',  300],
 ['Gym',    50],
)

# Start from the first cell below the headers.
row = 1
col = 0

# Iterate over the data and write it out row by row.
for item, cost in (expenses):
 worksheet.write(row, col,     item)
 worksheet.write(row, col + 1, cost, money)
 row += 1

# Write a total using a formula.
worksheet.write(row, 0, 'Total',       bold)
worksheet.write(row, 1, '=SUM(B2:B5)', money)

# Access the underlying xlsxwriter worksheet and write to it.
worksheet.write(0, 0, 'test')

# Apply conditional formatting
format1 = workbook.add_format({'bg_color':   '#FFC7CE',
                               'font_color': '#9C0006'})
worksheet.conditional_format('A1:Z1024', {'type':     'cell',
                                          'criteria': '>=',
                                          'value':    50,
                                          'format':   format1})

writer.save()

# Now save that as an image
import excel2img

# Save as PNG the range of used cells in test.xlsx on page named "Sheet1"
# excel2img.export_img("test.xlsx", "test.png", "Sheet1", None)

# Save as BMP the range B2:C15 in test.xlsx on page named "Sheet2"
excel2img.export_img("test.xlsx", "test.png", "", "Sheet1!A1:J30")


print("All Done")
# -----------------------------------------------------------------------------




# for col_range in range(1, 12):
#     cell_title = sheet.cell(1, col_range)
#     cell_title.fill = PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")
#
'''
# This takes around 141 sec
start = time.process_time()
xls = pd.ExcelFile('2019_04_01_AAII_DATA_Experiment.xlsm')
print(time.process_time() - start)

# This takes around 11 sec
start = time.process_time()
aaii_dateandperiod_df = pd.read_excel(xls, 'DateAndPeriod')
aaii_companyinfo_df = pd.read_excel(xls, 'CompanyInformation')
aaii_income_qtr_df = pd.read_excel(xls, 'IncomeSheet-QTR')
aaii_income_yr_df = pd.read_excel(xls, 'IncomeSheet-YR')
aaii_balancesheet_qtr_df = pd.read_excel(xls, 'BalanceSheet-QTR')
aaii_balancesheet_yr_df = pd.read_excel(xls, 'BalanceSheet-YR')
aaii_cashflow_qtr_df = pd.read_excel(xls, 'CashFlow-QTR')
aaii_cashflow_yr_df = pd.read_excel(xls, 'CashFlow-YR')
aaii_estimates_df = pd.read_excel(xls, 'EstimatesAndMultiples')
aaii_growth_df = pd.read_excel(xls, 'EstimatesAndMultiples')
aaii_misc_qtr_df = pd.read_excel(xls, 'Misc-QTR')
aaii_misc_yr_df = pd.read_excel(xls, 'Misc-YR')
aaii_rank_df = pd.read_excel(xls, 'Rank')
aaii_ratio_df = pd.read_excel(xls, 'RatiosAndValuations')
aaii_sector_df = pd.read_excel(xls, 'Median-Sectors')
aaii_industry_df = pd.read_excel(xls, 'Median-Industries')
print(time.process_time() - start)


ticker_list_unclean = aaii_dateandperiod_df['Ticker'].tolist()
ticker_list = [x for x in ticker_list_unclean if str(x) != 'nan']

for ticker_raw in ticker_list:

  ticker = ticker_raw.replace(" ", "").upper() # Remove all spaces from ticker_raw and convert to uppercase
  logging.info("========================================================")
  logging.info("Preparing 0 Ananlysis for " + ticker)
  logging.info("========================================================")

print ("All Done...")
'''


#
# # 6/30/2019 - This scripte works as inteneded but get thrown off when there are Yahoo and CNBC projections for the
# #  same date - an example is AAP - look at date 4/21/2017 this has Yahoo and CNBC data (In my opinion incorrectly
# # as Yahoo date is not the same row as the earnings is). Anyway because of the way the script is written it will
# # put the Yahoo date in the previous projections as all the rows what have nan in projection are deleted before
# # the processing so currently the scirpt has not way of knowing whether the Yahoo data is for the previous quater
# # or for the current quater. This shows up in the Earnings.csv file for AAP
# # One solution is to delted all the rows that have Yahoo is the column immpediately following projection column
# # that way we will lose that data but that should not be a big deal
#
# tracklist_df = pd.read_csv('Get_Earnings_Tracklist.csv')
# ticker_list_unclean = tracklist_df['Tickers'].tolist()
# ticker_list = [x for x in ticker_list_unclean if str(x) != 'nan']
#
# for ticker_raw in ticker_list:
#   ticker = ticker_raw.replace(" ", "").upper()  # Remove all spaces from ticker_raw and convert to uppercase
#   print("Getting Earnings for ", ticker)
#
#   # todo : Why does this not work here?
#   # earnings_df = pd.read_excel(dir_path + "\\" + "Stock_Files" + "\\" + ticker + '.xlsm', sheet_name="historical")
#   earnings_df = pd.read_excel('C:\Sundeep\Stocks_Automation\Scripts\Experiments\Stocks_Files' + "\\" + ticker + '.xlsm', sheet_name="historical")
#
#   print ("The Historical Tab (which contains earnings data) from the stock file is :", earnings_df)
#   print ("\nAll the columns in the historical tab are :\n")
#   for col in earnings_df.columns:
#     print(col)
#
#   # For some reason python reads the Date column as Timestamp
#   # Convert it to string here We will convet it to Datetime later
#   date_list = earnings_df['Date'].astype(str).tolist()
#   qtr_eps_list = earnings_df['Q EPS'].tolist()
#   projected_eps_list = earnings_df['projection'].tolist()
#   print ("The Raw Date list is :", date_list)
#
#   # Check if the length of all the columns are equal
#
#   # This works : The create a dataframe from a list of lists
#   step1_df=pd.DataFrame(list(zip(date_list, qtr_eps_list, projected_eps_list)),
#                         columns=['Date','Q EPS', 'projection'])
#   print ("Dataframe that only has Date, Q EPS and projection columns :", step1_df)
#
#   # ===========================================================================
#   # Clean up the dataframe
#   # ===========================================================================
#   # Remove all rows that do not have data in ALL columns
#   step1_df.dropna(how='all',inplace=True)
#   # Remove all rows that do not have data in Date column
#   step1_df.dropna(subset=['Date'],inplace=True)
#   tmp_df = step1_df[step1_df.Date != 'NaT']
#   step1_df = tmp_df
#   print ("New dataframe after dropping all rows with null in Date column:", step1_df)
#   # step1_df.to_csv('debug1.csv')
#   # Now we should have a dataframe that does not have any rows that do not have a valid date
#
#   # Drop ONLY the rows now that do not have any data in Q EPA AND projection column
#   step1_df.dropna(subset=['Q EPS', 'projection'], how='all', inplace=True)
#   # step1_df.to_csv('debug.csv')
#   print ("New dataframe after dropping all the rows with null in (Q EPS AND projection) columns:", step1_df)
#   # ===========================================================================
#
#   # ===========================================================================
#   # Extract the lists and write csv file
#   # ===========================================================================
#   csvFile=open('Extracted_Earnings' + "\\" + ticker + "_earnings.csv", 'w+', newline='')
#   writer = csv.writer(csvFile)
#   # Put the Header Row in the csv
#   writer.writerow(["Date", "Q_EPS_Diluted"])
#
#   step1_date_list = [dt.datetime.strptime(date, '%Y-%m-%d').date() for date in step1_df.Date.tolist()]
#   # step1_date_list = [dt.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').date() for date in step1_df.Date.tolist()]
#   step1_eps_list = step1_df['Q EPS'].tolist()
#   step1_projected_eps_list = step1_df['projection'].tolist()
#
#   tmp_index = 0
#   csv_line = []
#   for x in step1_date_list:
#     step1_eps = step1_eps_list[tmp_index]
#     step1_projected_eps = step1_projected_eps_list[tmp_index]
#     print ("Index : ", tmp_index, ", Date is : ", x, ", Q EPS is : ", step1_eps, ", Projected EPS is : ", step1_projected_eps)
#     if (str(step1_eps) != 'nan'):
#       print ("CSV String is :", csv_line)
#       if (tmp_index > 0):
#         writer.writerow(csv_line)
#
#       csv_line = []
#       print ("Found 1st Instance")
#       csv_line.insert(0, x.strftime('%m/%d/%Y'))
#       csv_line.insert(1, step1_eps)
#       csv_line.insert(2, step1_projected_eps)
#       insert_index = 3
#     else:
#       csv_line.insert(insert_index, step1_projected_eps)
#       insert_index = insert_index+1
#     tmp_index = tmp_index+1
#
# csvFile.close()
#
